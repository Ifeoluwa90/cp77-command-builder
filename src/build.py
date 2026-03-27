#!/usr/bin/env python3
"""
CP77 Command Builder — Build Script
Parses data/Cyberpunk_2077_Items.xlsx and generates index.html
Usage: python src/build.py
Requires: pip install openpyxl
"""

import json, os, re
from openpyxl import load_workbook

ROOT      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_FILE = os.path.join(ROOT, 'data', 'Cyberpunk_2077_Items.xlsx')
OUT_FILE  = os.path.join(ROOT, 'index.html')

_id_counter = 0
def make_id(cmd, name=''):
    global _id_counter
    _id_counter += 1
    m = re.search(r'"([^"]{3,})"', cmd)
    if m:
        return m.group(1)
    slug = re.sub(r'[^A-Za-z0-9_]', '_', name)[:50]
    return f'{slug}_{_id_counter}'

def clean(v):
    return str(v).strip() if v is not None else ''

def item(cmd, desc, sub, tier=None, notes=None):
    cmd = cmd.strip()
    parts = [p for p in [desc, tier, notes] if p and p not in ('-','None','none','')]
    full_desc = ' — '.join(parts) if parts else cmd[:80]
    return {'id': make_id(cmd, desc), 'cmd': cmd, 'desc': full_desc, 'sub': (sub or '').upper()}

def find_header(ws, keyword='command'):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, cell in enumerate(row):
            if cell and keyword in str(cell).lower():
                return i
    return None

def data_rows(ws, hdr_idx):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > hdr_idx and any(v is not None for v in row):
            yield row

# ── sheet parsers ────────────────────────────────────────────────────────────

def parse_standard(ws, label, sub_c, name_c, cmd_c, tier_c=None, notes_c=None):
    hi = find_header(ws)
    if hi is None:
        print(f'  [SKIP] {label} — no header found'); return []
    items, last_sub = [], ''
    for row in data_rows(ws, hi):
        def col(c): return clean(row[c]) if c is not None and c < len(row) else ''
        cmd = col(cmd_c)
        if not cmd: continue
        sub = col(sub_c) or last_sub
        last_sub = sub
        items.append(item(cmd, col(name_c), sub,
                          tier=col(tier_c) if tier_c is not None else None,
                          notes=col(notes_c) if notes_c is not None else None))
    print(f'  [{len(items):4d}] {label}')
    return items

def parse_vehicles(ws):
    items, hf = [], False
    for row in ws.iter_rows(values_only=True):
        if not hf:
            if row[0] and str(row[0]).lower() == 'menu': hf = True
            continue
        if clean(row[0]) != 'Unlock': continue
        cmd = clean(row[4]) if len(row) > 4 else ''
        if not cmd: continue
        mfr  = clean(row[2]); name = clean(row[3])
        cat  = clean(row[1]); note = clean(row[5]) if len(row) > 5 else ''
        items.append(item(cmd, f'{mfr} {name}'.strip(), cat, notes=note or None))
    print(f'  [{len(items):4d}] Vehicles')
    return items

def parse_progression(ws):
    items, hf, last_sub = [], False, ''
    for row in ws.iter_rows(values_only=True):
        if not hf:
            if row[2] and 'command' in str(row[2]).lower(): hf = True
            continue
        name = clean(row[1]); cmd = clean(row[2])
        misc = clean(row[3]) if len(row) > 3 else ''
        if not cmd:
            if name: last_sub = name
            continue
        items.append(item(cmd, name, last_sub, notes=misc or None))
    print(f'  [{len(items):4d}] Progression')
    return items

def parse_misc(ws):
    items, hf, last_sub = [], False, ''
    for row in ws.iter_rows(values_only=True):
        if len(row) < 3:
            continue
        if not hf:
            if row[2] and 'command' in str(row[2]).lower(): hf = True
            continue
        name = clean(row[1])
        cmd  = clean(row[2]) if len(row) > 2 else ''
        if not cmd:
            if name: last_sub = name
            continue
        items.append(item(cmd, name, last_sub))
    print(f'  [{len(items):4d}] Misc Commands')
    return items

# ── main ─────────────────────────────────────────────────────────────────────

def parse_xlsx(path):
    print(f'Parsing {os.path.basename(path)} ...')
    wb = load_workbook(path, read_only=True)
    def ws(n): return wb[n]

    sections = []
    def add(name, items):
        if items: sections.append({'name': name, 'items': items})

    add('Weapons',              parse_standard(ws('Weapons - All v2.2'),
                                    'Weapons', 0, 1, 3, tier_c=2, notes_c=5))
    add('Grenades',             parse_standard(ws('Weapon - Grenades v2.2'),
                                    'Grenades', 0, 1, 3, tier_c=2, notes_c=4))
    add('Weapon Mods',          parse_standard(ws('Weapon Mods - All v2.2'),
                                    'Weapon Mods', 0, 1, 3, tier_c=2, notes_c=4))
    add('Cyberware',            parse_standard(ws('Cyberware All v2.2'),
                                    'Cyberware', 0, 1, 3, tier_c=2, notes_c=5))
    add('Quickhacks',           parse_standard(ws('Cyberdeck QuickHacks v2.2'),
                                    'Quickhacks', 0, 1, 3, tier_c=2, notes_c=4))
    add('Clothing',             parse_standard(ws('Clothing - All v2.2'),
                                    'Clothing', 0, 1, 2, notes_c=3))
    add('Clothing Outfit Sets', parse_standard(ws('Clothing Outfit Sets v2.2'),
                                    'Clothing Outfit Sets', 0, 2, 3, notes_c=4))
    add('Crafting & Recipes',   parse_standard(ws('Crafting Mats & Recipes - All v'),
                                    'Crafting & Recipes', 0, 2, 4, tier_c=3, notes_c=5))
    add('Consumables',          parse_standard(ws('Consumables v2.2'),
                                    'Consumables', 0, 1, 2, notes_c=3))
    add('Vehicles',             parse_vehicles(ws('Vehicles - All v2.3')))
    add('Teleport Locations',   parse_standard(ws('Teleports Locations v2.2'),
                                    'Teleport Locations', 0, 1, 2, notes_c=5))
    add('Progression',          parse_progression(ws('Progression Items v2.2')))
    add('Misc Commands',        parse_misc(ws('Misc Commands v2.3')))

    total = sum(len(s['items']) for s in sections)
    print(f'\n  {len(sections)} sections — {total:,} total commands')
    return sections


def build_html(sections, out_path):
    data = json.dumps(sections, separators=(',',':'), ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CP2077 Command Builder</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#0f0f1e;color:#dde;min-height:100vh;display:flex;flex-direction:column}}
#topbar{{background:#090916;border-bottom:1px solid #2a2a4a;padding:10px 14px;display:flex;gap:10px;flex-wrap:wrap;align-items:center;flex-shrink:0}}
#topbar h1{{color:#00d4ff;font-size:17px;letter-spacing:.5px;white-space:nowrap}}
#topbar h1 span{{color:#ff4466}}
#search{{flex:1;min-width:200px;padding:7px 12px;border-radius:6px;border:1px solid #333;background:#1a1a30;color:#dde;font-size:13px}}
#search:focus{{outline:none;border-color:#00d4ff}}
#qtyWrap{{display:flex;align-items:center;gap:6px;font-size:13px;color:#888;white-space:nowrap}}
#qty{{width:70px;padding:6px 8px;border-radius:6px;border:1px solid #333;background:#1a1a30;color:#dde;font-size:13px;text-align:center}}
.btn{{padding:7px 13px;border-radius:6px;border:none;cursor:pointer;font-size:12px;font-weight:600;transition:background .1s,transform .08s}}
.btn:active{{transform:scale(.97)}}
.btn-cyan{{background:#00d4ff;color:#000}}.btn-cyan:hover{{background:#33ddff}}
.btn-green{{background:#00cc77;color:#000}}.btn-green:hover{{background:#00ee88}}
.btn-ghost{{background:transparent;border:1px solid #333;color:#999}}.btn-ghost:hover{{border-color:#ff4466;color:#ff4466}}
#selBadge{{background:#0d2a3a;color:#00d4ff;border-radius:10px;padding:3px 10px;font-size:12px;font-weight:600;white-space:nowrap}}
#main{{display:flex;flex:1;overflow:hidden;min-height:0}}
#left{{width:55%;border-right:1px solid #2a2a4a;overflow-y:auto}}
.sec-block{{}}
.sec-head{{display:flex;align-items:center;gap:8px;padding:8px 14px;cursor:pointer;background:#13132a;border-bottom:1px solid #1e1e3a;position:sticky;top:0;z-index:2;user-select:none}}
.sec-head:hover{{background:#1c1c38}}
.sec-arrow{{font-size:10px;color:#445;transition:transform .18s;flex-shrink:0}}
.sec-arrow.open{{transform:rotate(90deg);color:#00d4ff}}
.sec-title{{font-size:13px;font-weight:600;color:#bbd;flex:1}}
.sec-count{{font-size:11px;color:#445}}
.sec-body{{display:none}}.sec-body.open{{display:block}}
.sub-head{{padding:4px 14px;font-size:10px;font-weight:700;color:#446;letter-spacing:.7px;background:#0c0c1e;border-bottom:1px solid #141426}}
.item-row{{display:flex;align-items:flex-start;gap:9px;padding:6px 14px;border-bottom:1px solid #141426;cursor:pointer;transition:background .08s}}
.item-row:hover{{background:#181830}}
.item-row.checked{{background:#0b1e30;border-left:3px solid #00d4ff}}
.item-row input{{accent-color:#00d4ff;width:14px;height:14px;flex-shrink:0;margin-top:2px;cursor:pointer}}
.item-label{{flex:1;min-width:0}}
.item-name{{font-size:12px;color:#ccd;line-height:1.4}}
.item-cmd{{font-size:10px;color:#446;font-family:monospace;margin-top:1px;word-break:break-all;line-height:1.3}}
.no-results{{padding:40px;text-align:center;color:#334;font-size:13px}}
#right{{width:45%;display:flex;flex-direction:column;min-height:0}}
#rhead{{padding:9px 13px;background:#090916;border-bottom:1px solid #2a2a4a;display:flex;justify-content:space-between;align-items:center;flex-shrink:0}}
#rhead span{{font-size:12px;color:#556}}
#selList{{flex:1;overflow-y:auto;padding:8px;display:flex;flex-direction:column;gap:4px}}
.sel-item{{display:flex;align-items:center;gap:6px;background:#111128;border:1px solid #1e1e3e;border-radius:5px;padding:5px 8px}}
.sel-name{{font-size:11px;color:#99b;flex:1;word-break:break-all;line-height:1.3}}
.sel-qty{{font-size:11px;color:#00d4ff;font-weight:700;white-space:nowrap}}
.sel-rm{{background:none;border:none;color:#445;cursor:pointer;font-size:14px;padding:0 2px;line-height:1;flex-shrink:0}}.sel-rm:hover{{color:#ff4466}}
.empty-msg{{padding:30px 20px;text-align:center;color:#334;font-size:12px;line-height:2}}
#outWrap{{border-top:1px solid #2a2a4a;flex-shrink:0}}
#outHead{{padding:7px 12px;background:#090916;border-bottom:1px solid #141426;display:flex;justify-content:space-between;align-items:center}}
#outHead span{{font-size:11px;color:#446}}
#output{{width:100%;height:170px;background:#06060f;color:#00ff99;font-family:'Courier New',monospace;font-size:11px;padding:10px 12px;border:none;resize:none;outline:none;line-height:1.6}}
</style>
</head>
<body>
<div id="topbar">
  <h1>CP2077 <span>Command Builder</span></h1>
  <input id="search" type="text" placeholder="Search items, IDs, categories..." oninput="onSearch()">
  <div id="qtyWrap">Qty: <input id="qty" type="number" value="1" min="1" max="999999"></div>
  <span id="selBadge">0 selected</span>
  <button class="btn btn-ghost" onclick="deselectAll()">Clear</button>
  <button class="btn btn-cyan" onclick="generate()">Generate &#9654;</button>
</div>
<div id="main">
  <div id="left"></div>
  <div id="right">
    <div id="rhead">
      <span>Selected items</span>
      <div style="display:flex;gap:6px">
        <button class="btn btn-green" id="copyBtn" onclick="copyOut()">Copy all</button>
        <button class="btn btn-ghost" onclick="clearOut()">Clear output</button>
      </div>
    </div>
    <div id="selList">
      <div class="empty-msg">Check items on the left, then hit<br><strong style="color:#00d4ff">Generate &#9654;</strong> to build your command block.</div>
    </div>
    <div id="outWrap">
      <div id="outHead"><span id="cmdCount">0 commands ready</span></div>
      <textarea id="output" readonly placeholder="// Commands appear here&#10;// Paste into the CET console in-game"></textarea>
    </div>
  </div>
</div>
<script>
const S={data};
let sel={{}};
const left=document.getElementById('left');

function qty(){{return parseInt(document.getElementById('qty').value)||1;}}
function esc(s){{return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}}

function buildLeft(){{
  left.innerHTML='';
  S.forEach((sec,si)=>{{
    const block=document.createElement('div');
    block.className='sec-block';
    const head=document.createElement('div');
    head.className='sec-head';
    head.innerHTML=`<span class="sec-arrow" id="a${{si}}">&#9658;</span><span class="sec-title">${{esc(sec.name)}}</span><span class="sec-count">${{sec.items.length}}</span>`;
    head.onclick=()=>toggleSec(si);
    const body=document.createElement('div');
    body.className='sec-body';
    body.id='b'+si;
    renderItems(body,sec.items);
    block.appendChild(head);
    block.appendChild(body);
    left.appendChild(block);
  }});
}}

function renderItems(container,items){{
  container.innerHTML='';
  if(!items.length){{container.innerHTML='<div class="no-results">No items.</div>';return;}}
  let lastSub='';
  items.forEach(item=>{{
    if(item.sub&&item.sub!==lastSub){{
      const sh=document.createElement('div');
      sh.className='sub-head';
      sh.textContent=item.sub;
      container.appendChild(sh);
      lastSub=item.sub;
    }}
    const row=document.createElement('div');
    const chk=!!sel[item.id];
    row.className='item-row'+(chk?' checked':'');
    row.innerHTML=`<input type="checkbox" ${{chk?'checked':''}}><div class="item-label"><div class="item-name">${{esc(item.desc)}}</div><div class="item-cmd">${{esc(item.cmd)}}</div></div>`;
    const toggle=()=>toggleItem(item);
    row.onclick=toggle;
    row.querySelector('input').onclick=e=>{{e.stopPropagation();toggleItem(item);}};
    container.appendChild(row);
  }});
}}

function toggleSec(si){{
  const body=document.getElementById('b'+si);
  const arr=document.getElementById('a'+si);
  const open=body.classList.toggle('open');
  arr.classList.toggle('open',open);
}}

function toggleItem(item){{
  if(sel[item.id])delete sel[item.id];
  else sel[item.id]={{...item,selQty:qty()}};
  syncOpenSections();
  updateBadge();
  renderSelList();
}}

function syncOpenSections(){{
  S.forEach((sec,si)=>{{
    const body=document.getElementById('b'+si);
    if(body&&body.classList.contains('open'))renderItems(body,sec.items);
  }});
}}

function deselectAll(){{
  sel={{}};
  syncOpenSections();
  updateBadge();
  renderSelList();
}}

function updateBadge(){{
  document.getElementById('selBadge').textContent=Object.keys(sel).length+' selected';
}}

function renderSelList(){{
  const keys=Object.keys(sel);
  const panel=document.getElementById('selList');
  if(!keys.length){{
    panel.innerHTML='<div class="empty-msg">Check items on the left, then hit<br><strong style="color:#00d4ff">Generate &#9654;</strong> to build your command block.</div>';
    return;
  }}
  panel.innerHTML=keys.map(id=>{{
    const s=sel[id];
    const safe=id.replace(/\\\\/g,'\\\\\\\\').replace(/'/g,"\\\\'");
    return `<div class="sel-item"><span class="sel-name">${{esc(s.desc)}}</span><span class="sel-qty">x${{s.selQty}}</span><button class="sel-rm" onclick="removeItem('${{safe}}')" title="Remove">&#10005;</button></div>`;
  }}).join('');
}}

function removeItem(id){{
  delete sel[id];
  syncOpenSections();
  updateBadge();
  renderSelList();
}}

function generate(){{
  const keys=Object.keys(sel);
  if(!keys.length){{document.getElementById('output').value='// No items selected.';return;}}
  const lines=keys.map(id=>{{
    const s=sel[id];
    const m=s.cmd.match(/^Game\\.AddToInventory\\("([^"]+)",/);
    return m?`Game.AddToInventory("${{m[1]}}", ${{s.selQty}})`:s.cmd;
  }});
  document.getElementById('output').value=lines.join('\\n');
  document.getElementById('cmdCount').textContent=lines.length+' command'+(lines.length!==1?'s':'')+' ready';
}}

function copyOut(){{
  const v=document.getElementById('output').value;
  if(!v||v.startsWith('//'))return;
  navigator.clipboard.writeText(v).then(()=>{{
    const btn=document.getElementById('copyBtn');
    btn.textContent='Copied!';
    setTimeout(()=>btn.textContent='Copy all',2000);
  }});
}}

function clearOut(){{
  document.getElementById('output').value='';
  document.getElementById('cmdCount').textContent='0 commands ready';
}}

function onSearch(){{
  const q=document.getElementById('search').value.toLowerCase().trim();
  if(!q){{buildLeft();return;}}
  left.innerHTML='';
  let total=0;
  S.forEach((sec,si)=>{{
    const matched=sec.items.filter(i=>
      (i.desc||'').toLowerCase().includes(q)||
      (i.cmd||'').toLowerCase().includes(q)||
      (i.sub||'').toLowerCase().includes(q)||
      (i.id||'').toLowerCase().includes(q)
    );
    if(!matched.length)return;
    total+=matched.length;
    const block=document.createElement('div');
    block.className='sec-block';
    const head=document.createElement('div');
    head.className='sec-head';
    head.innerHTML=`<span class="sec-arrow open">&#9658;</span><span class="sec-title">${{esc(sec.name)}}</span><span class="sec-count">${{matched.length}} match${{matched.length!==1?'es':''}}</span>`;
    const body=document.createElement('div');
    body.className='sec-body open';
    renderItems(body,matched);
    block.appendChild(head);
    block.appendChild(body);
    left.appendChild(block);
  }});
  if(!total)left.innerHTML=`<div class="no-results" style="margin-top:80px">No results for &ldquo;${{esc(q)}}&rdquo;</div>`;
}}

buildLeft();
</script>
</body>
</html>"""

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    kb = os.path.getsize(out_path) // 1024
    print(f'Wrote {out_path} ({kb} KB)')


if __name__ == '__main__':
    sections = parse_xlsx(XLSX_FILE)
    build_html(sections, OUT_FILE)
    print('Done.')
